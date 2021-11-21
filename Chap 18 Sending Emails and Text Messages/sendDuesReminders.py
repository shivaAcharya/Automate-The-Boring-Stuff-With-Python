#! python3
# sendDuesReminders.py - Program to send reminder email to members with unpaid dues

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

#TODO: Check each member's payment status
unpaidMembers = {}
for i in range(2, sheet.max_rows+1):
    payment = sheet.cell(row=i, column=lastCol).value
    if latestMonth != 'paid':
        name = sheet.cell(row=i, column=1).value
        email = sheet.cell(row=1, column=2).value 
        unpaidMembers[name] = email

#TODO: Log In to email account.
smtpObj = smtplib.SMTP('smtp@gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('##########@example.com',  sys.argv[1])

#TODO: Send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid. \nDear %s, \nOur record shows that you haven't paid dues for %s. Please make this payment. Thank you!" %(latestMonth, name, latestMonth)
    print('Sending email to %s' %email)
    sendemailStatus = smtpObj.sendmail('##########@gmail.com', email, body)

    if sendemailStatus != {}:
        print('There was a problem sending email to %s: %s' %(email, sendemailStatus))
    smtpObj.quit()
