#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb['Sheet']


PRICE_UPDATE = {'Garlic' : 3.09,
                'Celery': 1.17,
                'Lemon': 2.19}


for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATE:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATE[produceName]

wb.save('updatedProduceSales.xlsx')
